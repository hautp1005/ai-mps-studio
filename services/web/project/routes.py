import configparser
import logging
import os
import json
from datetime import datetime
from subprocess import Popen, PIPE
from subprocess import check_output
import flask
import requests
import urllib3
import pandas as pd
import openpyxl
from time import sleep
from flask import (
    current_app as app,
    send_from_directory,
    request,
    redirect,
    url_for,
    render_template,
    abort, jsonify,
    send_file
)
from flask_oidc import OpenIDConnect
from urllib3.exceptions import InsecureRequestWarning
from werkzeug.utils import secure_filename

from .models import *


app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
logging.basicConfig(level=logging.DEBUG)
x_config = configparser.ConfigParser()
x_config.read('./project/configurations.cfg')

ALLOWED_EXTENSIONS = {'zip'}

s1 = u'ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚÝàáâãèéêìíòóôõùúýĂăĐđĨĩŨũƠơƯưẠạẢảẤấẦầẨẩẪẫẬậẮắẰằẲẳẴẵẶặẸẹẺẻẼẽẾếỀềỂểỄễỆệỈỉỊịỌọỎỏỐốỒồỔổỖỗỘộỚớỜờỞởỠỡỢợỤụỦủỨứỪừỬửỮữỰựỲỳỴỵỶỷỸỹ'
s0 = u'AAAAEEEIIOOOOUUYaaaaeeeiioooouuyAaDdIiUuOoUuAaAaAaAaAaAaAaAaAaAaAaAaEeEeEeEeEeEeEeEeIiIiOoOoOoOoOoOoOoOoOoOoOoOoUuUuUuUuUuUuUuYyYyYyYy'

KEYCLOAK_URI = x_config.get("keycloak", "KEYCLOAK_URI")
LOGOUT_URI = x_config.get("uri", "LOGOUT_URI")
REDIRECT_URI = x_config.get("uri", "REDIRECT_URI")
O_DOMAIN = x_config.get("domain", "O_DOMAIN")
SUPPER_EMAIL = x_config.get("domain", "SUPPER_EMAIL")
API_VER = "/api/v1.0"
http_error_200 = 'Success'
req_session = requests.Session()
req_session.verify = False
urllib3.disable_warnings()
x_headers = {'X-Api-Key': SUPPER_EMAIL}
app_center_headers = {'accept': "application/json", 'X-API-Token': "59abbb100e3aa366952a02cbdf5661e32f781f3e"}

oidc = OpenIDConnect(app=app, credentials_store=flask.session)

TESTCASE_FILE = ""
TESTCASE_CONVERT_FILE = ""
EXPORT_TESTCASE_FILE = ""
DATA_SAVE = {}

# Create a new workbook to write the output
output_workbook = openpyxl.Workbook()


def get_shell_script_output_using_communicate(filename):
    session = Popen([filename], stdout=PIPE, stderr=PIPE)
    stdout, stderr = session.communicate()
    if stderr:
        raise Exception("Error " + str(stderr))
    return stdout.decode('utf-8')


def get_shell_script_output_using_check_output(filename):
    stdout = check_output([filename]).decode('utf-8')
    return stdout


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def remove_accents(input_str):
    s = ''
    print(input_str.encode('utf-8'))
    for c in input_str:
        if c in s1:
            s += s0[s1.index(c)]
        else:
            s += c
    return s


def convert_workbook(user_id):
    # Open the Excel file
    workbook = openpyxl.load_workbook(TESTCASE_FILE)

    # Select the worksheet by name
    worksheet = workbook['Checklist_']

    # Select the first sheet of the output workbook
    output_sheet = output_workbook.active

    # Iterate over each row in the original worksheet
    for row in worksheet.iter_rows():
        # Create a new row in the output worksheet
        output_row = []
        for cell in row:
            # Check if the cell is within a merged cell range
            if cell.coordinate in [cell_range.coord for cell_range in worksheet.merged_cells.ranges]:
                # Get the merged cell value and add it to the output row
                merged_cell = worksheet.cell(*worksheet.merged_cells(cell.row, cell.column)[0].coord)
                output_row.append(merged_cell.value)
            else:
                # Add the cell value to the output row
                output_row.append(cell.value)

        # Write the output row to the output worksheet
        output_sheet.append(output_row)

    # Save the output workbook
    output_workbook.save(TESTCASE_CONVERT_FILE)
    print("Output workbook was succeed")


def put_workbook_chat_gpt(user_id, convert_output_file_path):
    global DATA_SAVE

    print("Starting to read file excel")
    # Read file excel
    df = pd.read_excel(convert_output_file_path)
    # Set display options to show merged cells
    pd.set_option('display.expand_frame_repr', False)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    col1 = df['Testcase description']
    col2 = df['Test to perform']

    # List the required formats
    count = 0
    na_arr = []
    for i in range(len(col1)):
        if pd.isnull(col1[i]) is False:
            na_arr += [i]
            count += 1

    # print(na_arr)
    # Create an objects Series from na_arr
    na_series = pd.Series(na_arr)
    count = 0
    na_series_i = 0
    start_na_series = 0
    outputed = ""
    # Export message file
    for c_desc_i in range(len(col1)):
        if pd.isnull(col1[c_desc_i]) is False:
            count += 1
            test_perform_arr = []
            for na_series_i in range(start_na_series, len(na_series)):
                try:
                    # print(f"{na_arr[na_series_i]} >> {na_arr[na_series_i + 1]}")
                    for k in range(na_arr[na_series_i], na_arr[na_series_i + 1]):
                        if pd.isnull(col2[k]) is False:
                            test_perform_arr += [col2[k]]

                    start_na_series += 1
                except BaseException:
                    try:
                        for j in range(max(na_arr), max(na_arr) * 10):
                            if pd.isnull(col2[j]) is False:
                                test_perform_arr += [col2[j]]

                    except BaseException:
                        print("")

                output = ""
                for i, step in enumerate(test_perform_arr):
                    output += f"{step}\n"

                count_try = 0
                for i in range(30):
                    try:
                        outputed = output
                        # Define a prompt for the chatbot
                        url = "https://api.openai.com/v1/chat/completions"

                        payload = json.dumps({
                            "model": "gpt-3.5-turbo",
                            "messages": [
                                {
                                    "role": "user",
                                    "content": f"Hãy bổ sung Test to perform cho Testcase description: {col1[c_desc_i]} chatgpt phản hồi bằng tiếng việt: " + outputed
                                }
                            ],
                            "temperature": 0.7
                        })
                        headers = {
                            'Content-Type': 'application/json',
                            'Authorization': 'Bearer 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'
                        }

                        response = requests.request("POST", url, headers=headers, data=payload)

                        data = json.loads(response.text)

                        # truy cập vào value của content
                        content = data["choices"][0]["message"]["content"]
                        #
                        test_steps = [content]

                        print(f"*************{col1[c_desc_i]}*********************")
                        print(test_steps)
                        count_try = 0
                        sleep(5)
                        break
                    except BaseException:
                        count_try += 1
                        print(f"fail {count_try}")
                        outputed = output
                        sleep(5)

                DATA_SAVE[col1[c_desc_i]] = outputed

                break
    return DATA_SAVE


def export_data_chatgpt(dt, export_chatgpt_file_path, sheet_name):
    wb = openpyxl.load_workbook(export_chatgpt_file_path)
    sheet = wb[sheet_name]

    # Khai báo biến row bằng 1 để lưu vào cột A
    row = 1

    # Lưu từng test case vào cột A và các giá trị tương ứng vào cột B
    for key, value in dt.items():
        sheet.cell(row=row, column=1, value=key)
        row += 1

        items = value.split('\n')
        for item in items:
            sheet.cell(row=row, column=2, value=item)
            row += 1

    # Lưu file Excel
    wb.save(export_chatgpt_file_path)


def get(url, headers=None):
    try:
        res = req_session.get(url, headers=headers)
        if res.status_code == 200:
            return res.json()
        else:
            return res
    except urllib3.exceptions.HTTPError as e:
        return e


def check_authorize():
    if not oidc.user_loggedin:
        return False
    oidc.get_access_token()
    return True


@app.route('/login', methods=['GET'])
@oidc.require_login
def login():
    if not oidc.user_loggedin:
        return logout()
    return redirect('/')


@app.route('/', methods=["GET", "POST"])
# @oidc.require_login
def index():
    print('Is Login :' + str(oidc.user_loggedin))
    global TESTCASE_FILE, TESTCASE_CONVERT_FILE, EXPORT_TESTCASE_FILE
    TESTCASE_FILE = os.path.join(app.config['TESTCASE_FILE']) + "/tc_hautp2.xlsx"
    TESTCASE_CONVERT_FILE = os.path.join(app.config['TESTCASE_CONVERT_FILE']) + "/tc_output_hautp2.xlsx"
    EXPORT_TESTCASE_FILE = os.path.join(app.config['EXPORT_TESTCASE_FILE']) + "/tc_output_chatgpt_hautp2.xlsx"
    # create a file output chatgpt
    output_workbook.save(EXPORT_TESTCASE_FILE)

    if request.method == 'POST':
        # upload_time = str(datetime.now())
        for file in request.files.getlist('file'):
            # generate a secure filename using the original filename
            secure_filename(file.filename)

            # Save the file with the new filename
            # file.save(os.path.join(app.config['TESTCASE_FOLDER'], f"tc_{upload_time}.xlsx"))
            file.save(TESTCASE_FILE)
            # convert workbook
            convert_workbook("hautp2")
            # # put workbook chatgpt
            put_wb_chatgpt = put_workbook_chat_gpt("hautp2", TESTCASE_CONVERT_FILE)
            # export file excel from chatgpt
            export_data_chatgpt(put_wb_chatgpt, EXPORT_TESTCASE_FILE, "Sheet")

            # download file
            return redirect(url_for('download_file'))
        return render_template("home.html", msg="Files uploaded successfully.")

    return render_template("home.html", msg="")

    # if not check_authorize():
    #     return redirect(url_for('login'))
    # else:
    #     print(g.oidc_id_token)
    #     email = str(g.oidc_id_token['email'])
    #     user_id = str(email).replace("@vng.com.vn", "")
    #     print(email)
    #     # get role user_id logged
    #     user_id_logged = str(g.oidc_id_token['email'])
    #     print(user_id_logged)
    #     get_role_user_id_logged = get_user_role(user_id_logged)
    #     print("get_role_user_id_logged is " + get_role_user_id_logged)
    #
    #     TESTCASE_FILE = os.path.join(app.config['TESTCASE_FILE']) + f"/tc_{user_id}.xlsx"
    #     TESTCASE_CONVERT_FILE = os.path.join(app.config['TESTCASE_CONVERT_FILE']) + f"/tc_output_{user_id}.xlsx"
    #     EXPORT_TESTCASE_FILE = os.path.join(app.config['EXPORT_TESTCASE_FILE']) + f"/tc_output_chatgpt_{user_id}.xlsx"
    #
    #     # Check user id is exits to add
    #     is_exists = db.session.query(UserTbl).filter(UserTbl.user_id == email).first()
    #     if is_exists is None:
    #         print("Add user " + email)
    #         if email == "hautp2@vng.com.vn":
    #             db.session.add(
    #                 UserTbl(user_id=email, password="", name=str(email).replace("@vng.com.vn", ""), role="supper",
    #                         is_active=True, description=None))
    #         else:
    #             db.session.add(
    #                 UserTbl(user_id=email, password="", name=str(email).replace("@vng.com.vn", ""), role="view",
    #                         is_active=True, description=None))
    #         db.session.commit()
    #
    #     if request.method == 'POST':
    #         # upload_time = str(datetime.now())
    #         for file in request.files.getlist('file'):
    #
    #             # generate a secure filename using the original filename
    #             secure_filename(file.filename)
    #
    #             # Save the file with the new filename
    #             # file.save(os.path.join(app.config['TESTCASE_FOLDER'], f"tc_{upload_time}.xlsx"))
    #             file.save(TESTCASE_FILE)
    #             # convert workbook
    #             convert_workbook(user_id)
    #             # put workbook chatgpt
    #             put_wb_chatgpt = put_workbook_chat_gpt(user_id, TESTCASE_CONVERT_FILE)
    #             # export file excel from chatgpt
    #             export_data_chatgpt(put_wb_chatgpt, EXPORT_TESTCASE_FILE, "Sheet")
    #
    #             # download file
    #             return redirect(url_for('download_file'))
    #         return render_template("home.html", msg="Files uploaded successfully.")
    #
    #     return render_template("home.html", msg="")


@app.route('/download_file', methods=['GET'])
def download_file():
    # Return a response with the file attached
    logging.info(os.path.join(app.config['EXPORT_TESTCASE_FILE']) + "/tc_output_chatgpt_hautp2.xlsx")
    return send_file(os.path.join(app.config['EXPORT_TESTCASE_FILE']) + "/tc_output_chatgpt_hautp2.xlsx", as_attachment=True)


@app.route('/upload_file', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Empty filename submitted'}), 400

    # Replace 'uploads' with a directory path to store the uploaded files
    file.save(os.path.join(app.config['TESTCASE_FILE']) + "/tc_hautp2.xlsx")

    return jsonify({'message': 'File uploaded successfully'}), 200


@app.route("/upload1", methods=["POST"])
def upload_file1():
    file = request.files["file"]
    filename = file.filename
    file.save(f"{os.getenv('APP_FOLDER')}/project/testcase/{filename}")
    logging.info(f"{os.getenv('APP_FOLDER')}/project/testcase/{filename}")
    return "File uploaded successfully!"


@app.route('/logout')
# @oidc.require_login
def logout():
    oidc.logout()
    flask.session.clear()
    return redirect(LOGOUT_URI + REDIRECT_URI)


@app.route('/health', methods=['GET'])
def health_check():
    return 'OK'


@app.route('/oidc_callback')
# @oidc.require_login
def oidc_callback():
    return redirect("/")


@app.route("/static/<path:filename>")
def staticfiles(filename):
    return send_from_directory(app.config["STATIC_FOLDER"], filename)


@app.route("/media/<path:filename>")
def mediafiles(filename):
    return send_from_directory(app.config["MEDIA_FOLDER"], filename)


@app.route('/validate', methods=["POST"])
def validate():
    if request.method == 'POST' and request.form['pass'] == '000':
        return redirect(url_for("success"))
    else:
        abort(401)


@app.route('/revoke-token')
def revoke_oidc_token():
    print('Is Login :' + str(oidc.user_loggedin))
    if oidc.user_loggedin is False:
        oidc.logout()
        flask.session.clear()
        return redirect(LOGOUT_URI + REDIRECT_URI)


@app.route(API_VER + '/chat/images/', methods=['POST', 'GET'])
def get_chat_images():
    headers = request.headers
    auth = headers.get("X-Api-Key")
    if auth == SUPPER_EMAIL:
        result = jsonify(deviceList=[""]), 200
    else:
        result = jsonify({"message": "ERROR: Unauthorized"}), 401
    return result
