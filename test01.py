import pandas as pd
import requests
import json
import openpyxl

# Set up authentication
OUTPUT_FOLDER = "/home/hautp2/Desktop/ai-mps-studio/services/web/project/testcase-output/tc_output_hautp2.xlsx"
OUTPUT_CONVERTED_FOLDER = "/home/hautp2/Desktop/ai-mps-studio/services/web/project/testcase-output/tc_output_converted_hautp2.xlsx"
# Mở workbook và chọn sheet cần thao tác
wb = openpyxl.load_workbook(OUTPUT_CONVERTED_FOLDER)
sheet = wb['Sheet']

data = {}


def read_workbook(user_id):
    global data

    print("Starting to read file excel")
    # Read file excel
    df = pd.read_excel(OUTPUT_FOLDER)
    # Set display options to show merged cells
    pd.set_option('display.expand_frame_repr', False)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    col1 = df['Testcase description']
    col2 = df['Test to perform']

    # List the required formats
    count = 0
    na_arr = []
    for i in range(len(col1)):
        if pd.isnull(col1[i]) is False:
            na_arr += [i]
            count += 1

    # print(na_arr)
    # Create an objects Series from na_arr
    na_series = pd.Series(na_arr)
    count = 0
    na_series_i = 0
    start_na_series = 0
    msg = ""

    # Export message file
    for c_desc_i in range(len(col1)):
        if pd.isnull(col1[c_desc_i]) is False:
            count += 1
            test_perform_arr = []
            for na_series_i in range(start_na_series, len(na_series)):
                try:
                    # print(f"{na_arr[na_series_i]} >> {na_arr[na_series_i + 1]}")
                    for k in range(na_arr[na_series_i], na_arr[na_series_i + 1]):
                        if pd.isnull(col2[k]) is False:
                            test_perform_arr += [col2[k]]

                    start_na_series += 1
                except BaseException:
                    try:
                        for j in range(max(na_arr), max(na_arr) * 10):
                            if pd.isnull(col2[j]) is False:
                                test_perform_arr += [col2[j]]

                    except BaseException:
                        print("")

                output = ""
                for i, step in enumerate(test_perform_arr):
                    output += f"{step}\n"

                # Define a prompt for the chatbot
                url = "https://api.openai.com/v1/chat/completions"

                payload = json.dumps({
                    "model": "gpt-3.5-turbo",
                    "messages": [
                        {
                            "role": "user",
                            "content": f"Hãy bổ sung Test to perform cho Testcase description: {col1[c_desc_i]} chatgpt phản hồi bằng tiếng việt: " + output
                        }
                    ],
                    "temperature": 0.7
                })
                headers = {
                    'Content-Type': 'application/json',
                    'Authorization': 'Bearer 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'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                data = json.loads(response.text)

                # truy cập vào value của content
                content = data["choices"][0]["message"]["content"]
                #
                test_steps = [content]

                print(f"*************{col1[c_desc_i]}*********************")
                print(test_steps)
                # test_steps_str = "\n".join(test_steps)
                # df = pd.DataFrame({'Testcase description': [test_desc], 'Test to perform': [test_steps_str]})
                # df = df.assign(**{'Test to perform': df['Test to perform'].str.split('\n')}).explode('Test to perform')

                # data[col1[c_desc_i]] = content

                break


read_workbook("user_id")

# # Khai báo biến row bằng 1 để lưu vào cột A
# row = 1
#
# # Vòng lặp lưu từng test case vào cột A và các giá trị tương ứng vào cột B
# for key, value in data.items():
#     sheet.cell(row=row, column=1, value=key)
#     row += 1
#
#     items = value.split('\n')
#     for item in items:
#         sheet.cell(row=row, column=2, value=item)
#         row += 1
#
# # Lưu file Excel
# wb.save(OUTPUT_CONVERTED_FOLDER)