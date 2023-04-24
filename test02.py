import openpyxl

OUTPUT_CONVERTED_FOLDER = "/home/hautp2/Desktop/ai-mps-studio/services/web/project/testcase-output/tc_output_converted_hautp3.xlsx"

# Mở workbook và chọn sheet cần thao tác
wb = openpyxl.load_workbook(OUTPUT_CONVERTED_FOLDER)
sheet = wb['Sheet']

# Dữ liệu cần lưu vào file Excel
data = {'TestCase 1': 'battle1\nbattle2\nbattle3',
        'TestCase 2': 'battle4\nbattle5\nbattle6',
        'TestCase 3': 'battle7\nbattle8\nbattle9'}

# Khai báo biến row bằng 1 để lưu vào cột A
row = 1

# Vòng lặp lưu từng test case vào cột A và các giá trị tương ứng vào cột B
for key, value in data.items():
    sheet.cell(row=row, column=1, value=key)
    row += 1

    items = value.split('\n')
    for item in items:
        sheet.cell(row=row, column=2, value=item)
        row += 1

# Lưu file Excel
wb.save(OUTPUT_CONVERTED_FOLDER)